# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook, Workbook
from Classes import Stop


# Load in the workbook
wb = load_workbook('./file.xlsx')

ws = wb[wb.sheetnames[0]]

stops = []

rows = tuple(ws.rows)

for row in rows[1:]:
    args = [cell.value for cell in row]
    stop = Stop(*args)
    stops.append(stop)

print("Loaded Excel as list of objects")

wb = Workbook()
destination_file_name = "output_excel.xlsx"
ws1 = wb.active
ws1.title = "Stops"

header_row = list((vars(stop).keys()))
ws1.append(header_row)

for i, stop in enumerate(stops):
    data_row = list((vars(stop).values()))
    ws1.append(data_row)
    # for k, val in enumerate(temp):
    #     ws1.cell(row=i+2, column=k+1).value = val
wb.save('test.xlsx')


if __name__ == '__main__':
    print("Hello")
